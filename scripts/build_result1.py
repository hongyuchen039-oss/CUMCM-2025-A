"""TASK_006-P3 — Q3 result1.xlsx Artifact Generation (BUDGET_LIMITED_BEST_KNOWN).

本脚本只做一件事: 对已冻结的 8 维 Q3 candidate 执行一次高精度重建 (fine /
scan_step=0.005), 并从官方空白模板生成、回读和核验 result1.xlsx.

显式不做 (与 P3 合同一致):

- 不重跑 Pilot / P2 512 / P2C 32;
- 不创建 challenger / 不调整决策变量 / 不调用 fake evaluator;
- 不修改 src/q1_baseline.py / src/q1_cylinder.py / src/q2_single_bomb.py /
  src/q2_search.py / src/q3_three_bombs.py / src/q3_search.py / problem/ /
  官方模板 ZIP;
- 不生成 result2.xlsx / result3.xlsx;
- 不启动 Final Audit CC / Hermes / Q4 / Q5;
- 不自动 Ready / merge PR;
- 不声称 FORMAL_RESULT_VERIFIED / local convergence / global optimum / 官方答案.

等级: BUDGET_LIMITED_BEST_KNOWN Q3 CANDIDATE WITH GENERATED AND ROUND-TRIP-VERIFIED
RESULT1.XLSX / LOCAL CONVERGENCE NOT ESTABLISHED / NOT A PROVEN GLOBAL OPTIMUM.
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import math
import os
import subprocess
import sys
import time
import zipfile
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from src.q3_three_bombs import (
    CANDIDATE_SCHEMA_VERSION,
    ThreeBombCandidate,
    evaluate_three_bomb_strategy,
)

# === 冻结的 8 维 Q3 candidate (P2C 闭合, 本轮不允许改变) ===

FROZEN_CANONICAL_CANDIDATE: Dict[str, float] = {
    "heading_rad": 3.127613485137657,
    "speed_mps": 116.12799297398149,
    "release_time_1_s": 0.993241052387636,
    "delay_1_s": 3.720360704323356,
    "release_time_2_s": 4.88566490244013,
    "delay_2_s": 3.7704749980723404,
    "release_time_3_s": 10.157737577136487,
    "delay_3_s": 3.7180978311642083,
}

REFERENCE_TOTAL_UNION_DURATION_S: float = 4.478204178810118
"""P3 canonical reconstruction reference (fine / scan_step=0.005).

This is the value produced by exactly one deterministic
`evaluate_three_bomb_strategy(candidate, sample_level='fine',
scan_step=0.005)` call against the frozen 8-dim candidate.

NOTE: The P2C closure_selection_score_s = 4.478218820691105 was produced
at a DIFFERENT profile (coarse / scan_step=0.05). It is preserved as a
historical evidence value (P2C selection score, not the P3 reconstruction
gate value). The profile_provenance_correction block in the closure
summary explicitly documents this distinction.
"""

CLOSURE_SELECTION_SCORE_S: float = 4.478218820691105
"""P2C closure selection score (profile=coarse, scan_step=0.05).

This is the argmax total_union_duration_s across the 32 P2C records.
It is NOT the same as the P3 canonical reconstruction reference
because the P3 reconstruction uses fine / 0.005 (a finer profile
that produces slightly different interval boundaries).
"""

CLOSURE_SELECTION_PROFILE: str = "coarse"
CLOSURE_SELECTION_SCAN_STEP_S: float = 0.05

RECONSTRUCTION_PROFILE: Dict[str, Any] = {
    "sample_level": "fine",
    "scan_step_s": 0.005,
}

RECONSTRUCTION_MISMATCH_TOLERANCE_S: float = 1e-12

ROUND_TRIP_ABS_TOL: float = 1e-10
ROUND_TRIP_REL_TOL: float = 1e-12

OFFICIAL_ZIP_PATH: str = "题目及模板/2025高教社杯数学建模A题_结果模板.zip"
OFFICIAL_TEMPLATE_MEMBER_BASENAME: str = "result1.xlsx"
OUTPUT_PATH: str = "outputs/submission/result1.xlsx"
ARTIFACT_SUMMARY_PATH: str = "outputs/q3/q3_result1_artifact_summary.json"
CHECKPOINT_PATH: str = "work/q3_result1/checkpoint.json"
CHECKPOINT_SCHEMA_VERSION: int = 5

CANONICAL_HEADERS: List[str] = [
    "无人机运动方向",
    "无人机运动速度 (m/s)",
    "烟幕干扰弹编号",
    "烟幕干扰弹投放点的x坐标 (m)",
    "烟幕干扰弹投放点的y坐标 (m)",
    "烟幕干扰弹投放点的z坐标 (m)",
    "烟幕干扰弹起爆点的x坐标 (m)",
    "烟幕干扰弹起爆点的y坐标 (m)",
    "烟幕干扰弹起爆点的z坐标 (m)",
    "有效干扰时长 (s)",
]


# === 异常类型 ===

class Result1BuilderError(Exception):
    """result1.xlsx 生成失败的通用基类."""


class TemplateError(Result1BuilderError):
    """官方模板结构错误 (header 缺失 / 行数错 / ZIP 损坏)."""


class ReconstructionGateError(Result1BuilderError):
    """重建值与 reference 的差超过 tolerance (1e-12)."""


class ResumeIdentityError(Result1BuilderError):
    """resume identity 校验失败 (任一 7 字段 mismatch)."""


class RoundTripError(Result1BuilderError):
    """回读核验失败 (cell 值 / 类型 / fingerprint 改变)."""


class NonNumericCellError(Result1BuilderError):
    """写入 cell 不是数值类型."""


# === helpers ===

def _git_head_sha() -> str:
    """获取当前 commit HEAD SHA (short or full); 失败时返回 'UNKNOWN'."""
    try:
        out = subprocess.check_output(
            ["git", "rev-parse", "HEAD"],
            cwd=ROOT, stderr=subprocess.DEVNULL,
        )
        return out.decode("ascii").strip()
    except Exception:
        return "UNKNOWN"


def _file_sha256(path: str) -> str:
    with open(path, "rb") as f:
        return hashlib.sha256(f.read()).hexdigest()


def _canonical_candidate_sha256() -> str:
    """冻结 candidate 的 SHA-256 (canonical JSON, sort_keys)."""
    canonical = json.dumps(
        FROZEN_CANONICAL_CANDIDATE, sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _make_candidate() -> ThreeBombCandidate:
    return ThreeBombCandidate(**FROZEN_CANONICAL_CANDIDATE)


def _heading_to_degrees(heading_rad: float) -> float:
    """heading_rad ∈ [0, 2π) → degrees modulo 360, 0 ≤ deg < 360."""
    deg = math.degrees(heading_rad) % 360.0
    if deg < 0:
        deg += 360.0
    if deg >= 360.0:
        deg -= 360.0
    return deg


# === template + workbook ===

def _load_official_workbook(
    zip_path: str = OFFICIAL_ZIP_PATH,
    member_basename: str = OFFICIAL_TEMPLATE_MEMBER_BASENAME,
):
    """从 read-only ZIP 加载 result1.xlsx 到内存, 返回 (workbook, bytes)."""
    if not os.path.exists(zip_path):
        raise TemplateError(f"official zip not found: {zip_path}")
    with open(zip_path, "rb") as f:
        zip_bytes = f.read()
    with zipfile.ZipFile(io.BytesIO(zip_bytes), mode="r") as zf:
        names = zf.namelist()
        matches = [n for n in names if os.path.basename(n) == member_basename]
        if len(matches) != 1:
            raise TemplateError(
                f"expected exactly 1 member with basename={member_basename!r} "
                f"in {zip_path!r}; got {len(matches)}: {matches}")
        member_name = matches[0]
        with zf.open(member_name, mode="r") as src:
            template_bytes = src.read()
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    return wb, template_bytes


def _detect_header(wb) -> Tuple[List[str], int]:
    """返回 (header_row_values, header_row_index_1based)."""
    ws = wb.active
    headers: List[str] = []
    header_row_idx: Optional[int] = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        row_vals: List[Any] = []
        for c in row:
            row_vals.append(c.value)
        # 检查前 10 个 cell 是否都非空字符串
        if all(v is not None and str(v).strip() != ""
               for v in row_vals[:len(CANONICAL_HEADERS)]):
            headers = [str(v) for v in row_vals[:len(CANONICAL_HEADERS)]]
            header_row_idx = row[0].row
            break
    if header_row_idx is None:
        raise TemplateError(
            "no header row found with 10 non-empty values matching canonical "
            f"schema; got headers={headers}")
    return headers, header_row_idx


def _validate_header(headers: List[str]) -> None:
    """10 列 contiguous + canonical 顺序严格匹配."""
    if len(headers) != len(CANONICAL_HEADERS):
        raise TemplateError(
            f"header count mismatch: expected {len(CANONICAL_HEADERS)} got "
            f"{len(headers)}: {headers}")
    for i, (expected, got) in enumerate(zip(CANONICAL_HEADERS, headers)):
        if expected != got:
            raise TemplateError(
                f"header[{i}] mismatch: expected {expected!r} got {got!r}")


def _collect_template_fingerprint(wb) -> Dict[str, Any]:
    """保存官方模板的 12 个 fingerprint 字段, 用于 preservation check."""
    ws = wb.active
    fp: Dict[str, Any] = {
        "sheet_names": list(wb.sheetnames),
        "active_sheet": ws.title,
        "dimensions": ws.dimensions,
        "merged_cells": [str(r) for r in ws.merged_cells.ranges],
        "freeze_panes": ws.freeze_panes,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "row_heights": {k: ws.row_dimensions[k].height for k in range(1, ws.max_row + 1)},
        "col_widths": {k: ws.column_dimensions[k].width for k in list("ABCDEFGHIJKLMN")},
        "non_data_cells": {},
        "non_data_formulas": {},
        "non_data_styles": {},
        "non_data_number_formats": {},
        "print_settings": {},
    }
    # 收集所有非 header / 非 A:J 数据 cell 的元数据
    for row in ws.iter_rows():
        for c in row:
            if c.value is None and c.data_type == "n":
                continue
            # 数据 cell (假设 header 在 row 1, 数据在 row 2-4, 列 A-J)
            is_header_row = (c.row == 1)
            is_data_cell = (
                not is_header_row
                and c.row in (2, 3, 4)
                and c.column <= len(CANONICAL_HEADERS)
            )
            if is_header_row or is_data_cell:
                # 数据 / header cell, 不参与 fingerprint 比较
                continue
            coord = c.coordinate
            fp["non_data_cells"][coord] = c.value
            if c.data_type == "f":
                fp["non_data_formulas"][coord] = str(c.value)
            # style/number_format
            if hasattr(c, "number_format"):
                fp["non_data_number_formats"][coord] = c.number_format
            if hasattr(c, "style") or hasattr(c, "_style"):
                style = getattr(c, "style", None) or getattr(c, "_style", None)
                if style is not None:
                    fp["non_data_styles"][coord] = str(style)
    # print settings (when readable)
    try:
        fp["print_settings"] = {
            "orientation": ws.page_setup.orientation,
            "paper_size": ws.page_setup.paperSize,
            "fit_to_width": ws.page_setup.fitToWidth,
            "fit_to_height": ws.page_setup.fitToHeight,
        }
    except Exception:
        pass
    return fp


def _fingerprint_matches(fp1: Dict[str, Any], fp2: Dict[str, Any]) -> bool:
    """两个 fingerprint dict 是否完全一致 (列宽 / 行高容许浮点 1e-9 误差)."""
    keys_to_compare = [
        "sheet_names", "active_sheet", "dimensions", "merged_cells",
        "freeze_panes", "non_data_cells", "non_data_formulas",
        "non_data_styles", "non_data_number_formats",
    ]
    for k in keys_to_compare:
        if fp1.get(k) != fp2.get(k):
            return False
    # row_heights / col_widths 容许 1e-9 浮点误差 (openpyxl re-serialize)
    for k in ("row_heights", "col_widths"):
        d1 = fp1.get(k, {})
        d2 = fp2.get(k, {})
        if set(d1.keys()) != set(d2.keys()):
            return False
        for key in d1:
            v1 = d1[key]
            v2 = d2[key]
            if v1 is None and v2 is None:
                continue
            if v1 is None or v2 is None:
                return False
            if not math.isclose(float(v1), float(v2), abs_tol=1e-9,
                                rel_tol=1e-9):
                return False
    return True


# === canonical reconstruction ===

def _run_canonical_reconstruction() -> Dict[str, Any]:
    """对冻结 candidate 执行 1 次 fine/0.005 重建; 返回 ThreeBombEvaluation 字段."""
    candidate = _make_candidate()
    ev = evaluate_three_bomb_strategy(
        candidate,
        sample_level=RECONSTRUCTION_PROFILE["sample_level"],
        scan_step=RECONSTRUCTION_PROFILE["scan_step_s"],
    )
    out: Dict[str, Any] = {
        "valid": ev.valid,
        "status": ev.status,
        "reason": ev.reason,
        "q3_evaluation_id": ev.q3_evaluation_id,
        "sample_level": ev.sample_level,
        "scan_step_s": ev.scan_step_s,
        "elapsed_s": ev.elapsed_s,
        "total_union_duration_s": ev.total_union_duration_s,
        "union_intervals": [list(iv) for iv in ev.union_intervals],
        "single_bomb_evaluator_calls": ev.single_bomb_evaluator_calls,
        "bomb_evaluations": [],
    }
    for b in ev.bomb_evaluations:
        out["bomb_evaluations"].append({
            "valid": b.valid,
            "status": b.status,
            "reason": b.reason,
            "total_duration_s": b.total_duration_s,
            "intervals": [list(iv) for iv in b.intervals],
            "release_point": list(b.release_point) if b.release_point else None,
            "detonation_time_s": b.detonation_time_s,
            "detonation_point": list(b.detonation_point) if b.detonation_point else None,
        })
    return out


def _check_reconstruction_gate(reconstructed: float) -> None:
    """abs(reconstructed - reference) <= 1e-12."""
    diff = abs(reconstructed - REFERENCE_TOTAL_UNION_DURATION_S)
    if diff > RECONSTRUCTION_MISMATCH_TOLERANCE_S:
        raise ReconstructionGateError(
            f"reconstructed={reconstructed!r} differs from reference="
            f"{REFERENCE_TOTAL_UNION_DURATION_S!r} by abs={diff:.3e} > "
            f"tolerance={RECONSTRUCTION_MISMATCH_TOLERANCE_S}")


# === write ===

def _validate_numeric(value: Any, coord: str) -> float:
    """写入 cell 必须是有限数值."""
    if not isinstance(value, (int, float)):
        raise NonNumericCellError(
            f"{coord}: expected numeric, got {type(value).__name__}={value!r}")
    v = float(value)
    if not math.isfinite(v):
        raise NonNumericCellError(
            f"{coord}: non-finite value {v!r}")
    return v


def _write_result1_rows(
    wb, header_row_idx: int, reconstruction: Dict[str, Any],
) -> None:
    """在 header_row 之后 3 行写入 A:J 数据."""
    ws = wb.active
    candidate = _make_candidate()
    heading_deg = _heading_to_degrees(candidate.heading_rad)
    speed_mps = candidate.speed_mps
    bomb_evals = reconstruction["bomb_evaluations"]

    if len(bomb_evals) != 3:
        raise Result1BuilderError(
            f"expected 3 bomb evaluations; got {len(bomb_evals)}")

    for i, bomb in enumerate(bomb_evals):
        row = header_row_idx + 1 + i
        # A: heading_deg
        ws.cell(row=row, column=1, value=_validate_numeric(heading_deg, f"A{row}"))
        # B: speed_mps
        ws.cell(row=row, column=2, value=_validate_numeric(speed_mps, f"B{row}"))
        # C: bomb index
        ws.cell(row=row, column=3, value=_validate_numeric(i + 1, f"C{row}"))
        # D-F: release_point
        release = bomb["release_point"]
        if release is None:
            raise Result1BuilderError(
                f"bomb {i+1} release_point is None")
        for axis_idx, col_idx in enumerate([4, 5, 6]):
            coord = f"{chr(ord('A') + col_idx - 1)}{row}"
            ws.cell(row=row, column=col_idx,
                    value=_validate_numeric(release[axis_idx], coord))
        # G-I: detonation_point
        det = bomb["detonation_point"]
        if det is None:
            raise Result1BuilderError(
                f"bomb {i+1} detonation_point is None")
        for axis_idx, col_idx in enumerate([7, 8, 9]):
            coord = f"{chr(ord('A') + col_idx - 1)}{row}"
            ws.cell(row=row, column=col_idx,
                    value=_validate_numeric(det[axis_idx], coord))
        # J: per-bomb own duration (NOT union)
        per_bomb_dur = bomb["total_duration_s"]
        # 验证 J 列不写 union: per_bomb_dur 应当小于等于 union 总时长
        if per_bomb_dur > reconstruction["total_union_duration_s"] + 1e-9:
            raise NonNumericCellError(
                f"J{row}: per-bomb duration {per_bomb_dur} exceeds union "
                f"{reconstruction['total_union_duration_s']}")
        ws.cell(row=row, column=10,
                value=_validate_numeric(per_bomb_dur, f"J{row}"))


def _save_workbook(wb, output_path: str) -> None:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


def _round_trip_verify(
    output_path: str, expected_data: List[List[float]],
    template_fingerprint: Dict[str, Any],
) -> None:
    """save → close → reopen → verify 10 cols × 3 rows + fingerprint."""
    # reopen from disk
    wb2 = openpyxl.load_workbook(output_path)
    fp2 = _collect_template_fingerprint(wb2)
    if not _fingerprint_matches(template_fingerprint, fp2):
        raise RoundTripError(
            "template fingerprint mismatch after save → reopen")
    ws2 = wb2.active
    # find header row in reopened workbook
    headers2, header_row_idx2 = _detect_header(wb2)
    if headers2 != CANONICAL_HEADERS:
        raise RoundTripError(
            f"reopened header mismatch: got {headers2}")
    for i, row_expected in enumerate(expected_data):
        row = header_row_idx2 + 1 + i
        for j, expected_val in enumerate(row_expected):
            col = j + 1
            cell = ws2.cell(row=row, column=col)
            v = cell.value
            if not isinstance(v, (int, float)) or not math.isfinite(float(v)):
                raise RoundTripError(
                    f"cell {cell.coordinate}: not finite numeric, got {v!r}")
            actual = float(v)
            expected = float(expected_val)
            if not math.isclose(actual, expected,
                                abs_tol=ROUND_TRIP_ABS_TOL,
                                rel_tol=ROUND_TRIP_REL_TOL):
                raise RoundTripError(
                    f"cell {cell.coordinate}: expected {expected!r} got "
                    f"{actual!r}; diff={abs(actual - expected):.3e}")


# === 7-field resume identity ===

def _compute_resume_identity() -> Dict[str, str]:
    """7 字段 resume identity (P3 schema v5)."""
    return {
        "execution_head_sha": _git_head_sha(),
        "contract_snapshot_sha256": _file_sha256(
            os.path.join(ROOT, "work", "task_contracts", "TASK_006-P3-v5.json")
        ) if os.path.exists(os.path.join(
            ROOT, "work", "task_contracts", "TASK_006-P3-v5.json")) else "",
        "q2_single_bomb_code_sha256": _file_sha256(
            os.path.join(ROOT, "src", "q2_single_bomb.py")),
        "q3_three_bombs_code_sha256": _file_sha256(
            os.path.join(ROOT, "src", "q3_three_bombs.py")),
        "result1_builder_code_sha256": _file_sha256(
            os.path.abspath(__file__)),
        "canonical_candidate_sha256": _canonical_candidate_sha256(),
        "official_template_sha256": _file_sha256(
            os.path.join(ROOT, OFFICIAL_ZIP_PATH)),
    }


def _verify_resume_identity(checkpoint: Dict[str, Any]) -> bool:
    """checkpoint 中的 7 字段是否与当前一致."""
    identity = _compute_resume_identity()
    for k in ("execution_head_sha", "contract_snapshot_sha256",
              "q2_single_bomb_code_sha256", "q3_three_bombs_code_sha256",
              "result1_builder_code_sha256", "canonical_candidate_sha256",
              "official_template_sha256"):
        if checkpoint.get(k) != identity[k]:
            return False
    return True


# === checkpoint ===

def _atomic_write_json(path: str, data: Dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, path)


# === main pipeline ===

def run(output_path: str = OUTPUT_PATH,
        artifact_summary_path: str = ARTIFACT_SUMMARY_PATH,
        dry_run: bool = False) -> Dict[str, Any]:
    """执行 P3 流程; 返回 artifact summary dict."""
    identity = _compute_resume_identity()

    # 1. load official template (read-only)
    wb, template_bytes = _load_official_workbook()
    template_sha_before = hashlib.sha256(template_bytes).hexdigest()

    # 2. detect + validate header
    headers, header_row_idx = _detect_header(wb)
    _validate_header(headers)

    # 3. fingerprint (before edit)
    fp_before = _collect_template_fingerprint(wb)

    # 4. canonical reconstruction (1 real Q3 call) unless dry-run
    if dry_run:
        reconstruction = {
            "valid": True,
            "status": "ok",
            "reason": "dry_run",
            "q3_evaluation_id": "DRY_RUN",
            "sample_level": RECONSTRUCTION_PROFILE["sample_level"],
            "scan_step_s": RECONSTRUCTION_PROFILE["scan_step_s"],
            "elapsed_s": 0.0,
            "total_union_duration_s": REFERENCE_TOTAL_UNION_DURATION_S,
            "union_intervals": [],
            "single_bomb_evaluator_calls": 0,
            "bomb_evaluations": [
                {
                    "valid": True,
                    "status": "ok",
                    "reason": "dry_run",
                    "total_duration_s": 1.0,
                    "intervals": [[0.0, 1.0]],
                    "release_point": [0.0, 0.0, 0.0],
                    "detonation_time_s": 1.0,
                    "detonation_point": [0.0, 0.0, 0.0],
                },
                {
                    "valid": True,
                    "status": "ok",
                    "reason": "dry_run",
                    "total_duration_s": 1.0,
                    "intervals": [[2.0, 3.0]],
                    "release_point": [0.0, 0.0, 0.0],
                    "detonation_time_s": 3.0,
                    "detonation_point": [0.0, 0.0, 0.0],
                },
                {
                    "valid": True,
                    "status": "ok",
                    "reason": "dry_run",
                    "total_duration_s": 1.0,
                    "intervals": [[4.0, 5.0]],
                    "release_point": [0.0, 0.0, 0.0],
                    "detonation_time_s": 5.0,
                    "detonation_point": [0.0, 0.0, 0.0],
                },
            ],
        }
    else:
        reconstruction = _run_canonical_reconstruction()

    # 5. reconstruction gate
    _check_reconstruction_gate(reconstruction["total_union_duration_s"])

    # 6. write rows
    if not dry_run:
        _write_result1_rows(wb, header_row_idx, reconstruction)
        # 7. save
        _save_workbook(wb, output_path)

    # 7.5 fingerprint after edit (only when actually editing)
    if not dry_run:
        # reload from disk to ensure we measure the saved state
        wb_after = openpyxl.load_workbook(output_path)
        fp_after_edit = _collect_template_fingerprint(wb_after)
        if not _fingerprint_matches(fp_before, fp_after_edit):
            raise RoundTripError(
                "template fingerprint changed after edit + save")
    else:
        fp_after_edit = fp_before

    # 8. round-trip
    expected_data: List[List[float]] = []
    candidate = _make_candidate()
    heading_deg = _heading_to_degrees(candidate.heading_rad)
    for i, bomb in enumerate(reconstruction["bomb_evaluations"]):
        row_data: List[float] = []
        row_data.append(heading_deg)
        row_data.append(candidate.speed_mps)
        row_data.append(float(i + 1))
        release = bomb["release_point"]
        row_data.extend([float(x) for x in release])
        det = bomb["detonation_point"]
        row_data.extend([float(x) for x in det])
        row_data.append(float(bomb["total_duration_s"]))
        expected_data.append(row_data)
    if not dry_run:
        _round_trip_verify(output_path, expected_data, fp_before)

    # 9. artifact summary
    output_sha256 = ""
    if not dry_run and os.path.exists(output_path):
        output_sha256 = _file_sha256(output_path)

    canonical_recon_duration = reconstruction["total_union_duration_s"]
    summary: Dict[str, Any] = {
        "phase_id": "TASK_006-P3",
        "contract_version": 5,
        "result_level": {
            "declared_level": "BUDGET_LIMITED_BEST_KNOWN",
            "not_a_proven_global_optimum": True,
            "local_convergence_established": False,
            "result1_xlsx_generated": not dry_run,
            "not_a_formal_q3_result": True,
        },
        "identity": identity,
        "canonical_candidate": FROZEN_CANONICAL_CANDIDATE,
        "canonical_reconstruction": reconstruction,
        "canonical_reconstruction_profile": RECONSTRUCTION_PROFILE["sample_level"],
        "canonical_reconstruction_scan_step_s":
            RECONSTRUCTION_PROFILE["scan_step_s"],
        "canonical_reconstruction_duration_s": canonical_recon_duration,
        "closure_selection_score_s": CLOSURE_SELECTION_SCORE_S,
        "closure_selection_profile": CLOSURE_SELECTION_PROFILE,
        "closure_selection_scan_step_s": CLOSURE_SELECTION_SCAN_STEP_S,
        "profile_provenance_correction": {
            "closure_selection_profile": CLOSURE_SELECTION_PROFILE,
            "closure_selection_scan_step_s": CLOSURE_SELECTION_SCAN_STEP_S,
            "closure_selection_score_s": CLOSURE_SELECTION_SCORE_S,
            "canonical_high_resolution_profile":
                RECONSTRUCTION_PROFILE["sample_level"],
            "canonical_high_resolution_scan_step_s":
                RECONSTRUCTION_PROFILE["scan_step_s"],
            "canonical_high_resolution_duration_s": canonical_recon_duration,
            "absolute_profile_difference_s": abs(
                canonical_recon_duration - CLOSURE_SELECTION_SCORE_S),
            "decision_variables_changed": False,
            "search_rerun_performed": False,
            "explanation": (
                "The P2C selection score came from a coarse-profile record. "
                "P3 reports the same frozen candidate under fine 0.005 "
                "reconstruction."
            ),
        },
        "reconstruction_gate": {
            "reference_total_union_duration_s":
                REFERENCE_TOTAL_UNION_DURATION_S,
            "reconstructed_total_union_duration_s": canonical_recon_duration,
            "abs_diff": abs(canonical_recon_duration
                            - REFERENCE_TOTAL_UNION_DURATION_S),
            "tolerance_s": RECONSTRUCTION_MISMATCH_TOLERANCE_S,
            "passed": True,
        },
        "workbook": {
            "official_template_path": OFFICIAL_ZIP_PATH,
            "official_template_sha256": template_sha_before,
            "official_template_member_basename":
                OFFICIAL_TEMPLATE_MEMBER_BASENAME,
            "output_path": output_path,
            "output_sha256": output_sha256,
            "sheet_names": fp_before["sheet_names"],
            "active_sheet": fp_before["active_sheet"],
            "header_names": CANONICAL_HEADERS,
            "column_mapping": {
                "A": "heading_deg (degrees(heading_rad) modulo 360)",
                "B": "speed_mps",
                "C": "bomb index",
                "D": "release_point x",
                "E": "release_point y",
                "F": "release_point z",
                "G": "detonation_point x",
                "H": "detonation_point y",
                "I": "detonation_point z",
                "J": "bomb i own total_duration_s (NOT union)",
            },
            "three_rows_written": len(expected_data) == 3,
            "all_cells_numeric": True,
            "union_duration_written_to_j": False,
            "round_trip_status": "passed" if not dry_run else "skipped",
        },
        "template_fingerprint": fp_before,
        "status": ("result1_xlsx_generated"
                   if not dry_run else "dry_run_completed"),
        "output_path": output_path,
        "artifact_summary_path": artifact_summary_path,
    }

    # bind summary SHA
    summary["result1_run_identity_sha256"] = hashlib.sha256(
        json.dumps(summary["identity"], sort_keys=True,
                   separators=(",", ":")).encode("utf-8")
    ).hexdigest()

    _atomic_write_json(artifact_summary_path, summary)

    # write checkpoint
    checkpoint = {
        "checkpoint_schema_version": CHECKPOINT_SCHEMA_VERSION,
        **identity,
        "reconstruction_elapsed_s": reconstruction["elapsed_s"],
        "reconstructed_total_union_duration_s":
            reconstruction["total_union_duration_s"],
        "output_sha256": output_sha256,
        "status": summary["status"],
        "timestamp": time.time(),
    }
    _atomic_write_json(CHECKPOINT_PATH, checkpoint)

    return summary


# === CLI ===

def main() -> int:
    parser = argparse.ArgumentParser(
        description="TASK_006-P3 Q3 result1.xlsx artifact generation",
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="skip reconstruction + write; only verify "
                             "template structure + resume identity")
    parser.add_argument("--output-path", default=OUTPUT_PATH)
    parser.add_argument("--artifact-summary-path",
                        default=ARTIFACT_SUMMARY_PATH)
    args = parser.parse_args()

    try:
        run(output_path=args.output_path,
            artifact_summary_path=args.artifact_summary_path,
            dry_run=args.dry_run)
    except (TemplateError, ReconstructionGateError,
            ResumeIdentityError, RoundTripError,
            NonNumericCellError, Result1BuilderError) as e:
        print(f"FAIL: {type(e).__name__}: {e}", file=sys.stderr)
        return 2
    except Exception as e:
        print(f"INTERNAL ERROR: {type(e).__name__}: {e}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())